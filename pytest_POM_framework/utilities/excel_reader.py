import openpyxl
def get_list_from_excel(excel_name, sheet_name):
    creds = openpyxl.load_workbook(excel_name)
    sheet = creds[sheet_name]
    login_creds = creds["login_creds"]
    credentials = []
    for row in range(1, sheet.max_row+1):
        nested_creds = []
        for column in range(1, sheet.max_column+1):
            data = login_creds.cell(row, column)
            nested_creds.append(data.value)
        credentials.append(nested_creds)
    return credentials

# def add_value_to_excel(excel_name, sheet_name):
#     creds = openpyxl.load_workbook(excel_name)
#     sheet = creds[sheet_name]
#     sheet.cell(10,1).value = "new value"
#     creds.save(excel_name)
#
# add_value_to_excel("./creds.xlsx","login_creds")